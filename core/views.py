import json
import string
import secrets
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q, Count

from .models import User, Cabinet, Phase, File
from .forms import (
    LoginForm, CabinetForm, PhaseForm, FileForm,
    UserCreateForm, UserEditForm,
)
from .decorators import admin_required


# ─── Authentication ─────────────────────────────────────────────
def login_view(request):
    if request.user.is_authenticated:
        return redirect('core:dashboard')
    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            messages.success(request, f'Welcome back, {user.first_name}!')
            return redirect('core:dashboard')
        else:
            messages.error(request, 'Invalid email or password.')
    else:
        form = LoginForm()
    return render(request, 'core/login.html', {'form': form})


def logout_view(request):
    logout(request)
    messages.info(request, 'You have been logged out.')
    return redirect('core:login')


# ─── Dashboard ──────────────────────────────────────────────────
STATUS_LABELS = dict(File.STATUS_CHOICES)

@login_required
def dashboard(request):
    total_files = File.objects.filter(is_deleted=False).count()
    total_cabinets = Cabinet.objects.count()
    total_users = User.objects.filter(is_active=True).count()
    total_creations = File.objects.count()
    latest_files = File.objects.filter(is_deleted=False).select_related(
        'cabinet', 'created_by'
    )[:10]

    # File status breakdown for chart
    status_qs = (
        File.objects.filter(is_deleted=False)
        .values('status')
        .annotate(count=Count('id'))
        .order_by('status')
    )
    chart_labels = []
    chart_data = []
    for entry in status_qs:
        chart_labels.append(STATUS_LABELS.get(entry['status'], entry['status']))
        chart_data.append(entry['count'])

    context = {
        'total_files': total_files,
        'total_cabinets': total_cabinets,
        'total_users': total_users,
        'total_creations': total_creations,
        'latest_files': latest_files,
        'chart_labels': json.dumps(chart_labels),
        'chart_data': json.dumps(chart_data),
    }
    return render(request, 'core/dashboard.html', context)


# ─── Cabinet Management ────────────────────────────────────────
@login_required
def cabinet_list(request):
    from django.db.models import Prefetch
    cabinets = Cabinet.objects.prefetch_related(
        Prefetch('phases', queryset=Phase.objects.prefetch_related('files').order_by('order')),
    ).all()
    return render(request, 'core/cabinet_list.html', {'cabinets': cabinets})


@login_required
def cabinet_create(request):
    if request.method == 'POST':
        form = CabinetForm(request.POST)
        if form.is_valid():
            cabinet = form.save(commit=False)
            cabinet.created_by = request.user
            cabinet.save()
            messages.success(request, f'Cabinet "{cabinet.name}" created successfully!')
            return redirect('core:cabinet_list')
    else:
        form = CabinetForm()
    return render(request, 'core/cabinet_create.html', {'form': form})


@login_required
def cabinet_detail(request, pk):
    cabinet = get_object_or_404(Cabinet, pk=pk)
    phases = cabinet.phases.all()
    files = cabinet.files.filter(is_deleted=False)
    return render(request, 'core/cabinet_detail.html', {
        'cabinet': cabinet,
        'phases': phases,
        'files': files,
    })


@login_required
def cabinet_edit(request, pk):
    cabinet = get_object_or_404(Cabinet, pk=pk)
    if request.method == 'POST':
        form = CabinetForm(request.POST, instance=cabinet)
        if form.is_valid():
            form.save()
            messages.success(request, f'Cabinet "{cabinet.name}" updated successfully!')
            return redirect('core:cabinet_detail', pk=pk)
    else:
        form = CabinetForm(instance=cabinet)
    return render(request, 'core/cabinet_create.html', {
        'form': form,
        'editing': True,
        'cabinet': cabinet,
    })


@login_required
def cabinet_delete(request, pk):
    cabinet = get_object_or_404(Cabinet, pk=pk)
    if request.method == 'POST':
        name = cabinet.name
        cabinet.delete()
        messages.success(request, f'Cabinet "{name}" deleted successfully!')
        return redirect('core:cabinet_list')
    return render(request, 'core/delete_confirm.html', {
        'object': cabinet,
        'object_type': 'Cabinet',
        'cancel_url': 'core:cabinet_list',
    })


# ─── Phase Management ──────────────────────────────────────────
@login_required
def phase_create(request, cabinet_pk):
    cabinet = get_object_or_404(Cabinet, pk=cabinet_pk)
    if request.method == 'POST':
        form = PhaseForm(request.POST)
        if form.is_valid():
            phase = form.save(commit=False)
            phase.cabinet = cabinet
            phase.created_by = request.user
            phase.save()
            messages.success(request, f'Phase "{phase.name}" added to {cabinet.name}!')
            return redirect('core:cabinet_detail', pk=cabinet_pk)
    else:
        # Suggest next phase number
        existing_count = cabinet.phases.count()
        form = PhaseForm(initial={
            'name': f'Phase {existing_count + 1}',
            'order': existing_count + 1,
        })
    return render(request, 'core/phase_create.html', {
        'form': form,
        'cabinet': cabinet,
    })


@login_required
def phase_delete(request, pk):
    phase = get_object_or_404(Phase, pk=pk)
    cabinet_pk = phase.cabinet.pk
    if request.method == 'POST':
        name = phase.name
        phase.delete()
        messages.success(request, f'Phase "{name}" deleted successfully!')
        return redirect('core:cabinet_detail', pk=cabinet_pk)
    return render(request, 'core/delete_confirm.html', {
        'object': phase,
        'object_type': 'Phase',
        'cancel_url': 'core:cabinet_detail',
        'cancel_pk': cabinet_pk,
    })


# ─── File Management ───────────────────────────────────────────
@login_required
def file_list(request):
    files = File.objects.filter(is_deleted=False).select_related(
        'cabinet', 'phase', 'created_by'
    )
    # Search
    q = request.GET.get('q', '')
    if q:
        files = files.filter(
            Q(file_name__icontains=q) |
            Q(file_number__icontains=q) |
            Q(cabinet__name__icontains=q)
        )
    # Filter by cabinet
    cabinet_id = request.GET.get('cabinet', '')
    if cabinet_id:
        files = files.filter(cabinet_id=cabinet_id)
    # Filter by status
    status = request.GET.get('status', '')
    if status:
        files = files.filter(status=status)

    cabinets = Cabinet.objects.all()
    return render(request, 'core/file_list.html', {
        'files': files,
        'cabinets': cabinets,
        'q': q,
        'selected_cabinet': cabinet_id,
        'selected_status': status,
    })


@login_required
def file_create(request):
    if request.method == 'POST':
        form = FileForm(request.POST, request.FILES)
        if form.is_valid():
            file_obj = form.save(commit=False)
            file_obj.created_by = request.user
            file_obj.save()
            messages.success(request, f'File "{file_obj.file_name}" created successfully!')
            return redirect('core:file_list')
    else:
        form = FileForm()
    cabinets = Cabinet.objects.all()
    return render(request, 'core/file_create.html', {
        'form': form,
        'cabinets': cabinets,
    })


@login_required
def file_detail(request, pk):
    file_obj = get_object_or_404(File, pk=pk, is_deleted=False)
    return render(request, 'core/file_detail.html', {'file': file_obj})


@login_required
def file_edit(request, pk):
    file_obj = get_object_or_404(File, pk=pk, is_deleted=False)
    if request.method == 'POST':
        form = FileForm(request.POST, request.FILES, instance=file_obj)
        if form.is_valid():
            file_obj = form.save(commit=False)
            file_obj.updated_by = request.user
            file_obj.save()
            messages.success(request, f'File "{file_obj.file_name}" updated successfully!')
            return redirect('core:file_detail', pk=pk)
    else:
        form = FileForm(instance=file_obj)
    return render(request, 'core/file_edit.html', {
        'form': form,
        'file': file_obj,
    })


@login_required
def file_delete(request, pk):
    file_obj = get_object_or_404(File, pk=pk, is_deleted=False)
    if request.method == 'POST':
        file_obj.is_deleted = True
        file_obj.deleted_by = request.user
        file_obj.deleted_at = timezone.now()
        file_obj.save()
        messages.success(request, f'File "{file_obj.file_name}" deleted successfully!')
        return redirect('core:file_list')
    return render(request, 'core/delete_confirm.html', {
        'object': file_obj,
        'object_type': 'File',
        'cancel_url': 'core:file_list',
    })


# ─── User Management (Admin Only) ──────────────────────────────
@login_required
@admin_required
def user_list(request):
    users = User.objects.all().order_by('-date_joined')
    return render(request, 'core/user_list.html', {'users': users})


@login_required
@admin_required
def user_create(request):
    generated_password = None
    if request.method == 'POST':
        form = UserCreateForm(request.POST)
        if form.is_valid():
            user = form.save()
            generated_password = form.cleaned_data['password']
            messages.success(
                request,
                f'User {user.email} created successfully! '
                f'Password: {generated_password}'
            )
            return redirect('core:user_list')
    else:
        # Auto-generate a strong password suggestion
        generated_password = ''.join(
            secrets.choice(string.ascii_letters + string.digits)
            for _ in range(10)
        )
        form = UserCreateForm(initial={'password': generated_password})
    return render(request, 'core/user_create.html', {
        'form': form,
        'generated_password': generated_password,
    })


@login_required
@admin_required
def user_edit(request, pk):
    user_obj = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=user_obj)
        if form.is_valid():
            form.save()
            messages.success(request, f'User {user_obj.email} updated successfully!')
            return redirect('core:user_list')
    else:
        form = UserEditForm(instance=user_obj)
    return render(request, 'core/user_edit.html', {
        'form': form,
        'user_obj': user_obj,
    })


@login_required
@admin_required
def user_toggle_active(request, pk):
    user_obj = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        user_obj.is_active = not user_obj.is_active
        user_obj.save()
        status = 'activated' if user_obj.is_active else 'deactivated'
        messages.success(request, f'User {user_obj.email} {status}.')
    return redirect('core:user_list')


# ─── AJAX API Endpoints ────────────────────────────────────────
@login_required
def api_file_number_suggestions(request):
    query = request.GET.get('q', '')
    if len(query) < 2:
        return JsonResponse({'suggestions': []})
    suggestions = File.objects.filter(
        file_number__icontains=query, is_deleted=False
    ).values_list('file_number', flat=True)[:10]
    return JsonResponse({'suggestions': list(suggestions)})


@login_required
def api_check_file_number(request):
    number = request.GET.get('number', '')
    exists = File.objects.filter(file_number=number, is_deleted=False).exists()
    return JsonResponse({'exists': exists})


@login_required
def api_phases(request):
    cabinet_id = request.GET.get('cabinet_id', '')
    if not cabinet_id:
        return JsonResponse({'phases': []})
    phases = Phase.objects.filter(cabinet_id=cabinet_id).values('id', 'name')
    return JsonResponse({'phases': list(phases)})


# ─── Reports ───────────────────────────────────────────────────
@login_required
def reports_page(request):
    status_choices = File.STATUS_CHOICES
    return render(request, 'core/reports.html', {
        'status_choices': status_choices,
    })


@login_required
def generate_report(request, report_type):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse

    files = File.objects.filter(is_deleted=False).select_related(
        'cabinet', 'phase', 'created_by'
    )

    if report_type == 'all_files':
        title = 'All Files Report'
        filename = 'EGIS_All_Files_Report.xlsx'
    else:
        # Filter by specific status
        files = files.filter(status=report_type)
        label = STATUS_LABELS.get(report_type, report_type)
        title = f'{label} Report'
        filename = f'EGIS_{report_type}_Report.xlsx'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Report'

    # Styles
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # Title row
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = Font(name='Calibri', bold=True, size=16, color='4F46E5')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36

    # Subtitle row
    ws.merge_cells('A2:G2')
    sub_cell = ws['A2']
    sub_cell.value = f'Generated on {timezone.now().strftime("%B %d, %Y at %I:%M %p")}'
    sub_cell.font = Font(name='Calibri', italic=True, size=10, color='64748B')
    sub_cell.alignment = Alignment(horizontal='center')

    # Headers
    headers = ['S/N', 'File Name', 'File Number', 'Case Brief', 'Status', 'Date Invited', 'Created By', 'Created At']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    for idx, f in enumerate(files, 1):
        row_num = idx + 4
        date_inv = ''
        if f.date_invited:
            date_inv = f.date_invited.strftime('%a %d %B, %Y')
        row_data = [
            idx,
            f.file_name,
            f.file_number,
            f.case_brief or '',
            f.display_status,
            date_inv,
            str(f.created_by) if f.created_by else 'Unknown',
            f.created_at.strftime('%Y-%m-%d %H:%M'),
        ]
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    # Auto-width columns
    col_widths = [6, 20, 18, 65, 22, 22, 20, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def generate_pdf_report(request, report_type):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from django.http import HttpResponse
    import io

    files = File.objects.filter(is_deleted=False).select_related(
        'cabinet', 'phase', 'created_by'
    )

    if report_type == 'all_files':
        title = 'All Files Report'
        filename = 'EGIS_All_Files_Report.pdf'
    else:
        files = files.filter(status=report_type)
        label = STATUS_LABELS.get(report_type, report_type)
        title = f'{label} Report'
        filename = f'EGIS_{report_type}_Report.pdf'

    # ── Column colors (harmonious palette) ──────────────────
    col_header_colors = [
        colors.HexColor('#4F46E5'),   # Indigo   — S/N
        colors.HexColor('#0D9488'),   # Teal     — File Name
        colors.HexColor('#D97706'),   # Amber    — File Number
        colors.HexColor('#7C3AED'),   # Violet   — Case Brief
        colors.HexColor('#059669'),   # Emerald  — Status
        colors.HexColor('#0284C7'),   # Sky      — Date Invited
        colors.HexColor('#E11D48'),   # Rose     — Created By
        colors.HexColor('#475569'),   # Slate    — Created At
    ]

    col_light_colors = [
        colors.HexColor('#EEF2FF'),   # Indigo light
        colors.HexColor('#F0FDFA'),   # Teal light
        colors.HexColor('#FFFBEB'),   # Amber light
        colors.HexColor('#F5F3FF'),   # Violet light
        colors.HexColor('#ECFDF5'),   # Emerald light
        colors.HexColor('#F0F9FF'),   # Sky light
        colors.HexColor('#FFF1F2'),   # Rose light
        colors.HexColor('#F8FAFC'),   # Slate light
    ]

    # ── Build PDF ───────────────────────────────────────────
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Title'],
        fontSize=18,
        textColor=colors.HexColor('#4F46E5'),
        spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        'ReportSubtitle',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.HexColor('#64748B'),
        spaceAfter=14,
    )
    cell_style = ParagraphStyle(
        'CellText',
        parent=styles['Normal'],
        fontSize=8,
        leading=10,
    )
    header_style = ParagraphStyle(
        'HeaderText',
        parent=styles['Normal'],
        fontSize=9,
        leading=11,
        textColor=colors.white,
        fontName='Helvetica-Bold',
    )

    elements = []

    # Title & subtitle
    elements.append(Paragraph(title, title_style))
    elements.append(Paragraph(
        f'Generated on {timezone.now().strftime("%B %d, %Y at %I:%M %p")}',
        subtitle_style,
    ))

    # Table headers
    headers = ['S/N', 'File Name', 'File Number', 'Case Brief', 'Status',
               'Date Invited', 'Created By', 'Created At']
    header_row = [Paragraph(h, header_style) for h in headers]

    # Table data
    data = [header_row]
    for idx, f in enumerate(files, 1):
        date_inv = ''
        if f.date_invited:
            date_inv = f.date_invited.strftime('%a %d %b, %Y')
            
        # Clean and truncate fields to prevent PDF LayoutErrors from massive text or many newlines
        def clean_text(text, max_len):
            if not text: return ''
            cleaned = str(text).replace('\n', ' ').replace('\r', ' ').strip()
            # Escape XML characters for ReportLab Paragraph
            cleaned = cleaned.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            if len(cleaned) > max_len:
                return cleaned[:max_len-3] + '...'
            return cleaned

        row = [
            Paragraph(str(idx), cell_style),
            Paragraph(clean_text(f.file_name, 100), cell_style),
            Paragraph(clean_text(f.file_number, 50), cell_style),
            Paragraph(clean_text(f.case_brief, 150), cell_style),
            Paragraph(clean_text(f.display_status, 50), cell_style),
            Paragraph(date_inv, cell_style),
            Paragraph(clean_text(f.created_by if f.created_by else 'Unknown', 50), cell_style),
            Paragraph(f.created_at.strftime('%Y-%m-%d %H:%M'), cell_style),
        ]
        data.append(row)

    # Column widths (landscape A4 ≈ 277mm usable)
    col_widths = [30, 80, 55, 120, 60, 62, 55, 50]

    table = Table(data, colWidths=col_widths, repeatRows=1)

    # ── Table style ─────────────────────────────────────────
    style_commands = [
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
    ]

    # Per-column header colors
    for col_idx, hdr_color in enumerate(col_header_colors):
        style_commands.append(
            ('BACKGROUND', (col_idx, 0), (col_idx, 0), hdr_color)
        )

    # Per-column alternating row colors
    num_rows = len(data)
    for row_idx in range(1, num_rows):
        for col_idx, light_color in enumerate(col_light_colors):
            if row_idx % 2 == 0:
                style_commands.append(
                    ('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), light_color)
                )
            else:
                style_commands.append(
                    ('BACKGROUND', (col_idx, row_idx), (col_idx, row_idx), colors.white)
                )

    table.setStyle(TableStyle(style_commands))
    elements.append(table)

    doc.build(elements)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/pdf',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    buffer.close()
    return response

